import pandas as pd
from rdflib import URIRef, BNode, Literal, Graph
from rdflib.namespace import RDF, RDFS, FOAF, XSD
from rdflib import Namespace
import numpy as np
import math
import sys
import argparse
import json
import html

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import openpyxl

def value(ws, j, c):
    return ws.cell(row=j+1, column=c+1).value

def getH(value):
    if type(value) is int:
        return [value]
    else:
        arr = value.split(",")
        arr2 = []
        for e in arr:
            arr2.append(int(e))
        return arr2
    

def read_excel(path):

    '''
    df = pd.read_excel(path, sheet_name=0, header=None, index_col=None, engine="openpyxl")

    r_count = len(df.index)
    c_count = len(df.columns)
    '''

    wb = load_workbook(path)
    ws = wb.active

    r_count = ws.max_row
    c_count = ws.max_column

    map = {}

    for i in range(0, c_count):
        # label = df.iloc[1, i]
        label = ws.cell(row=1+1, column=i+1).value
        map[i] = label

    print(map)

    data = []
    # search = []

    uri_context = "https://static.toyobunko-lab.jp/taishozo/context.json"

    context = {
        "@context" : [
            {
                "ex" : "http://example.org/",
                "data" : "https://static.toyobunko-lab.jp/taishozo/data",
                "keiten" : "https://static.toyobunko-lab.jp/taishozo/経典番号/"
            }
        ]
    }

    with open("../static/context.json", 'w') as f:
        json.dump(context, f, ensure_ascii=False, indent=4,
        sort_keys=True, separators=(',', ': '))

    tmp_arr = []

    for j in range(2, r_count):
        # id = df.iloc[j, 0]
        id = ws.cell(row=j+1, column=0+1).value

        if id == "" or id == None: # pd.isnull(id):
            continue

        id = str(id).zfill(5)

        print(id)

        if id == "00000":
            continue
        
        '''
        if id != "04309":
            continue
        '''

        '''
        if j > 10:
            break
        '''

        # 経典番号 = df.iloc[j, 1]
        経典番号 = ws.cell(row=j+1, column=1+1).value

        uri = "data:"+id+".json"
        uri_経典番号 = "keiten:"+経典番号+".json"

        # 枝番 = df.iloc[j, 9] if not pd.isnull(df.iloc[j, 9]) else ""
        枝番 = ws.cell(row=j+1, column=9+1).value if ws.cell(row=j+1, column=9+1).value else ""

        # print(df.iloc[j, 8], df.iloc[j, 9], df.iloc[j, 10], df.iloc[j, 11], df.iloc[j, 12], df.iloc[j, 13])

       
        # uri_sat = "https://21dzk.l.u-tokyo.ac.jp/SAT2018/"+df.iloc[j, 8]+枝番+"_."+str(df.iloc[j, 10]).zfill(2)+"."+str(df.iloc[j, 11]).zfill(4)+df.iloc[j, 12]+str(df.iloc[j, 13]).zfill(2)+".html"

        # --------

        texts_k = []

        for c in range(0, 5):

            start = c * 10

            c22 = value(ws, j, 22 + start) # ws.cell(row=j+1, column=22+start+1).value
            c23 = value(ws, j, 23 + start)
            c24 = value(ws, j, 24 + start)

            c25 = value(ws, j, 25 + start)
            c26 = value(ws, j, 26 + start)
            c27 = value(ws, j, 27 + start)
            c28 = value(ws, j, 28 + start)
            c29 = value(ws, j, 29 + start)
            c30 = value(ws, j, 30 + start)
            c31 = value(ws, j, 31 + start)


            if c22:
                obj = {
                    "@id": uri+"#テキスト"+str(c+1)+"（勘同目録）",
                    "ex:標準名称" : c22,
                }
                if c23:
                    obj["ex:巻"] = c23
                if c24:
                    obj["ex:国"] = c24
                if c25:
                    obj["ex:時代"] = c25
                if c26:
                    obj["ex:年"] = c26
                if c27:
                    obj["ex:～年"] = c27
                if c28:
                    obj["ex:刊写者"] = c28

                if c29:
                    obj["ex:刊写形態"] = c29

                if c30:
                    obj["ex:関与者"] = c30

                if c31:
                    obj["ex:関与形態"] = c31

                texts_k.append(obj)

        

        # --------

        holds_k = []

        for c in range(0, 2):

            start = c * 2

            c72 = value(ws, j, 72 + start)
            c73 = value(ws, j, 73 + start)

            if c72:
                obj = {
                    "@id": uri+"#所蔵者"+str(c+1)+"（勘同目録）",
                    "ex:国" : c72
                }
                if c73:
                    obj["ex:所蔵者"] = c73
                holds_k.append(obj)

        # --------

        c14 = value(ws, j, 14)
        c15 = value(ws, j, 15)
        c16 = value(ws, j, 16)
        c17 = value(ws, j, 17)

        obj_k = {
            "@id" : uri+"#勘同目録",
            "ex:texts" : texts_k,
            "ex:所蔵者" : holds_k
        }

        if c14:
            obj_k["ex:底本/校本"] = c14

        if c15:
            obj_k["ex:❹"] = c15

        if c16:
            obj_k["ex:❼"] = c16

        if c17:
            obj_k["ex:❼備考"] = c17

        # --------

        

        texts_f = []
  
        for c in range(0, 3):

            start = c * 10

            c79 = value(ws, j, 79+start)
            c80 = value(ws, j, 80+start)
            c81 = value(ws, j, 81+start)
            c82 = value(ws, j, 82+start)
            c83 = value(ws, j, 83+start)
            c84 = value(ws, j, 84+start)
            c85 = value(ws, j, 85+start)
            c86 = value(ws, j, 86+start)
            c87 = value(ws, j, 87+start)
            c88 = value(ws, j, 88+start)


            if c79:
                obj = {
                    "@id": uri+"#テキスト"+str(c+1)+"（脚注）",
                    "ex:標準名称" : c79,
                }
                if c80:
                    obj["ex:巻"] = c80

                if c81:
                    obj["ex:国"] = c81
                if c82:
                    obj["ex:時代"] = c82
                if c83:
                    obj["ex:年"] = c83
                if c84:
                    obj["ex:～年"] = c84
                if c85:
                    obj["ex:刊写者"] = c85

                if c86:
                    obj["ex:刊写形態"] = c86

                if c87:
                    obj["ex:関与者"] = c87

                if c88:
                    obj["ex:関与形態"] = c88

                texts_f.append(obj)
       

        # --------

        

        holds_f = []

        for c in range(0, 2):

            start = c * 2

            ci = value(ws, j, 109 + start)
            ci2 = value(ws, j, 110 + start)

            if ci:
                obj = {
                    "@id": uri+"#所蔵者"+str(c+1)+"（脚注）",
                    "ex:国" : ci,
                }
                
                if ci2:
                    obj["ex:所蔵者"] = ci2
                holds_f.append(obj)

        # --------

        c18 = value(ws, j, 18)
        c19 = value(ws, j, 19)
        c20 = value(ws, j, 20)
        c21 = value(ws, j, 21)
        c76 = value(ws, j, 76)
        c77 = value(ws, j, 77)
        c78 = value(ws, j, 78)

        obj_f = {
            "@id" : uri+"#脚注",
            "ex:texts" : texts_f,
            "ex:所蔵者" : holds_f
        }

        if c18:
            obj_f["ex:底本/校本"] = c18

        if c19:
            obj_f["ex:新添"] = c19

        if c20:
            obj_f["ex:テキスト"] = c20

        if c21:
            obj_f["ex:備考"] = c21

        if c76:
            obj_f["ex:底本推定"] = c76

        if c77:
            obj_f["ex:略号の使用"] = c77

        if c78:
            obj_f["ex:略号解説"] = c78

        # --------

        

        経典名 = ws.cell(row=j+1, column=3+1).value
        

        spl = 経典名.split("\"")

        # 経典名 = spl[3]
        uri_sat = spl[1]

        # print(経典名, uri_sat)        

        keiten = {
            "@id" : uri_経典番号,
            "ex:経典番号" : 経典番号,
            "ex:枝番" : value(ws, j, 2),
            "ex:経典名" : spl[3],
            "ex:収録巻次" : value(ws, j, 4),
            "ex:部門" : value(ws, j, 5),
            "ex:配本" : getH(value(ws, j, 6)),
            # "ex:url":  spl[1]
            
        }

        date = str(value(ws, j, 7)).split(",")
        keiten["ex:出版年月日"] = []

        for d in date:
            d = str(d)
            keiten["ex:出版年月日"].append(d[0:4]+"-"+d[4:6]+"-"+d[6:8])

        # --------
        # 勘同目録IIIFコレクション

        spl = value(ws, j, 122).split("\"")

        obj_ki = {
            "@id" : uri+"#勘同目録IIIFコレクション",
            "ex:kmID" : value(ws, j, 121),
            "ex:ページ" : spl[3],
            "ex:段" : value(ws, j, 123),
            "ex:toページ" : value(ws, j, 124),
            "ex:to段" : value(ws, j, 125),
            "ex:url": spl[1]
        }

        # --------
        # 酉目

        yus = []

        for c in range(0, 2):

            start = c * 4
            
            index = 113 + start
            if value(ws, j, index):
                spl = value(ws, j, index + 1).split("\"")
                obj = {
                    "@id": uri+"#酉目"+str(c+1)+"",
                    "ex:經典番號" : value(ws, j, index),
                    "ex:通番" : spl[3],
                    "ex:枝番" : value(ws, j, index + 2),
                    "ex:to枝番" : value(ws, j, index + 3),
                    "ex:url" : spl[1]
                }
                yus.append(obj)

        # --------
        # 酉蓮社本IIIコレクション

        yu_i = []

        for c in range(0, 3):
            start = c * 4
            index = 126 + start
            if value(ws, j, index):

                spl = value(ws, j, index).split("\"")

                obj = {
                    "@id": uri+"#酉蓮社本IIIFコレクション"+str(c+1)+"",
                    "ex:経典名" : spl[3],
                    "ex:画像フォルダ" : value(ws, j, index + 1),
                    "ex:from" : value(ws, j, index + 2),
                    "ex:to" : value(ws, j, index + 3),
                    "ex:url" : spl[1]
                }
                yu_i.append(obj)
        # --------

        '''
        keiten = ""
        obj_k = ""
        obj_f = ""
        obj_ki = ""
        yus = ""
        yu_i = ""
        '''
        
        obj = {
            "@id" : uri,
            "@context": uri_context,
            "ex:基本情報" : [
                {
                    "@id" : uri+"#基本情報",
                    "ex:No." : id,
                    "ex:経典" : [
                        keiten
                    ]
                }
            ],
            "ex:sat" : [
                {
                    "@id" : uri+"#SAT頭出し用",
                    "ex:url" : uri_sat,
                }
            ],
            "ex:勘同目録" : [
                obj_k
            ],
            "ex:脚注" : [
                obj_f
            ],
            "ex:勘同目録IIIFコレクション" : [
                obj_ki
            ],
            "ex:酉目" : yus,
            "ex:酉蓮社本IIIコレクション" : yu_i
        }

        '''
        # ゆうれんじゃ本の経典名(２)が存在するもの
        if value(ws, j, 130):
            # print(df.iloc[j, 0], df.iloc[j, 126])
            tmp_arr.append(value(ws, j, 0))
        '''

        # --------

        data.append(obj)

        with open("../static/data/"+id+".json", 'w') as f:
            json.dump(obj, f, ensure_ascii=False, indent=4,
            sort_keys=True, separators=(',', ': '))

        # ------------

    with open("../static/data.json", 'w') as f:
        json.dump(data, f, ensure_ascii=False, indent=4,
        sort_keys=True, separators=(',', ': '))

   

    print(tmp_arr)
    print(len(tmp_arr))


path = "../static/metadata/data.xlsx"
# data1 = read_excel(path)
read_excel(path)