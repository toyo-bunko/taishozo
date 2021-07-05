import pandas as pd
from rdflib import URIRef, BNode, Literal, Graph
from rdflib.namespace import RDF, RDFS, FOAF, XSD
from rdflib import Namespace
import numpy as np
import math
import sys
import argparse
import json
import html
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import openpyxl



with open("/Users/nakamurasatoru/git/d_sat/u-renja/static/iiif2/collection/top.json") as f:
    df2 = json.load(f)

manifests = df2["manifests"]

images = {}

for m in manifests:
    metadata = m["metadata"]

    num = -1
    identifier = ""

    for obj in metadata:
        if "Description" in obj["label"]:
            num = obj["value"][0].split("通番: ")[1]

            num = int(num)

        elif "identifier" in obj["label"]:
            identifier = obj["value"]

    if num not in images:
        images[num] = []

    images[num].append({
        "id": m["@id"],
        "identifier": identifier
    })

'''
for num in images:
    print(num, images[num])
'''

#############

url = "https://static.toyobunko-lab.jp/taishozo/iiif/kandomokuroku/manifest.json"

df = requests.get(url).json()

indexMap = {}

canvases = df["sequences"][0]["canvases"]
for i in range(len(canvases)):
    c = canvases[i]
    res = c["images"][0]["resource"]["@id"]
    if "p." in res:
        p = res.split("p.")[1].split(".")[0]
        indexMap[p] = i + 1

'''
for p in indexMap:
    print(p, indexMap[p], int(p) - int(indexMap[p]))
'''

hyperlink = Font(underline='single', color='0563C1')

def read_excel(path):
    # df = pd.read_excel(path, sheet_name=0, header=None, index_col=None, engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active

    r_count = ws.max_row
    c_count = ws.max_column

    map = {}

    for i in range(0, c_count):
        label = ws.cell(row=1+1, column=i+1).value
        map[i] = label

    for j in range(2, r_count):
        id = ws.cell(row=j+1, column=0+1).value

        if id == "" or id == None:
            continue

        経典番号 =  ws.cell(row=j+1, column=1+1).value
        枝番 = ws.cell(row=j+1, column=9+1).value
        if 枝番 == None:
            枝番 = ""

        e1 = ws.cell(row=j+1, column=8+1).value if ws.cell(row=j+1, column=8+1).value != None else ws.cell(row=j+1, column=8+1).value # ""
        e2 = ws.cell(row=j+1, column=10+1).value if ws.cell(row=j+1, column=10+1).value != None else ws.cell(row=j+1, column=10+1).value # ""
        e3 = ws.cell(row=j+1, column=11+1).value if ws.cell(row=j+1, column=11+1).value != None else ws.cell(row=j+1, column=11+1).value # ""
        e4 = ws.cell(row=j+1, column=12+1).value if ws.cell(row=j+1, column=12+1).value != None else ws.cell(row=j+1, column=12+1).value # ""
        e5 = ws.cell(row=j+1, column=13+1).value if ws.cell(row=j+1, column=13+1).value != None else ws.cell(row=j+1, column=13+1).value # ""

        uri_sat = "https://21dzk.l.u-tokyo.ac.jp/SAT2018/"+e1+枝番+"_."+str(e2).zfill(2)+"."+str(e3).zfill(4)+e4+str(e5).zfill(2)+".html"

        title = ws.cell(row=j+1, column=3+1).value
        
        ws.cell(row=j+1, column=3+1).value = '=HYPERLINK("{}", "{}")'.format(uri_sat, title)
        ws.cell(row=j+1, column=3+1).font = hyperlink
        
        num1 = ws.cell(row=j+1, column=114+1).value #df.iloc[j, 114]
        if num1 != "" and num1 != None:
            ws.cell(row=j+1, column=114+1).value = '=HYPERLINK("https://static.toyobunko-lab.jp/u-renja/search/?main[refinementList][通番]={}", "{}")'.format(num1, num1)
            ws.cell(row=j+1, column=114+1).font = hyperlink
            
            # df.iloc[j, 114] = "[{}]({})".format(num1, num1)

        num2 = ws.cell(row=j+1, column=118+1).value #df.iloc[j, 114]
        if num2 != "" and num2 != None:
            ws.cell(row=j+1, column=118+1).value = '=HYPERLINK("https://static.toyobunko-lab.jp/u-renja/search/?main[refinementList][通番]={}", "{}")'.format(str(num2).zfill(4), num2)
            ws.cell(row=j+1, column=118+1).font = hyperlink

        kando = ws.cell(row=j+1, column=122+1).value #df.iloc[j, 114]
        if kando != "" and kando != None:
            pos = int(kando) - 152
            ws.cell(row=j+1, column=122+1).value = '=HYPERLINK("http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest=https://static.toyobunko-lab.jp/taishozo/iiif/kandomokuroku/manifest.json&pos={}", "{}")'.format(pos, int(kando))
            ws.cell(row=j+1, column=122+1).font = hyperlink
        
        # 画像リンク
        for c in range(0, 3):

            

            folder1 = ws.cell(row=j+1, column=127+1 + 4 * c).value #df.iloc[j, 114]
            
            
            if folder1 != "" and folder1 != None:

                x = ws.cell(row=j+1, column=128+1 + 4 * c).value
                y = ws.cell(row=j+1, column=129+1 + 4 * c).value

                # print(folder1, x, y)

                uuid1 = folder1 + "_" + str(x).zfill(4) + "_" + str(y).zfill(4)
                
                num1 = int(num1)

                num = num1

                if uuid1 == "u-renja1524-1525_0003_0005":
                    num += 1

                if uuid1 == "u-renja1663-1668_0015_0017":
                    num += 1

                if uuid1 == "u-renja1663-1668_0017_0019":
                    num += 2

                if num in images:

                    value1 = ""
                        
                    arr = images[num]

                    for a in arr:
                        if uuid1 == a["identifier"]:
                            value1 = a["id"]

                    if value1 == "":
                        print("missing", uuid1, "value1", value1, "num", num)
                    else:
                        ws.cell(row=j+1, column=126+1 + 4 * c).value = '=HYPERLINK("http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest={}", "{}")'.format(value1, ws.cell(row=j+1, column=126+1 + 4 * c).value)
                        ws.cell(row=j+1, column=126+1 + 4 * c).font = hyperlink
                        # df.iloc[j, 126] = "[{}]({})".format(df.iloc[j, 126], value1)

    wb.save('../static/metadata/data.xlsx')


with open("config.json") as f:
    config = json.load(f)

filename = config["filename"]

path = "data/" + filename

# data1 = read_excel(path)
read_excel(path)